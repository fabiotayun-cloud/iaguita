import hashlib
import json
import re
import traceback
from datetime import datetime
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_file
import face_recognition
from PIL import Image, ImageDraw, ImageFont
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XlImage

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024  # 32MB max

BASE_DIR = Path(__file__).parent
MEMBERS_DIR = BASE_DIR / 'miembros'
TRAINING_DIR = BASE_DIR / 'entrenamientos'
INDEX_PATH = TRAINING_DIR / 'index.json'
EXCEL_PATH = BASE_DIR / 'asistencias.xlsx'
DEMO_DIR = BASE_DIR / 'demo_fotos_grupales'
DEMOS_DIR = BASE_DIR / 'demos'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp', 'bmp'}

# Cache global de miembros: {nombre: encoding}
members_cache = {}


# ─────────────────────────────────────────────
# Utilidades generales
# ─────────────────────────────────────────────

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def load_image_from_upload(file_storage):
    """Carga una imagen desde un FileStorage de Flask y la convierte a formato RGB numpy array."""
    img = Image.open(file_storage)
    img = img.convert('RGB')
    return np.array(img)


def detect_faces_multiscale(image, upsample=2):
    """
    Detecta rostros usando HOG con múltiples escalas para encontrar caras pequeñas.
    Escala la imagen hacia arriba en varias pasadas y combina resultados,
    eliminando detecciones duplicadas.
    """
    all_locations = set()

    # Pasada 1: imagen original con upsample configurado
    locs = face_recognition.face_locations(image, number_of_times_to_upsample=upsample, model='hog')
    for loc in locs:
        all_locations.add(loc)

    # Pasada 2: imagen escalada x1.5 (para caras pequeñas)
    h, w = image.shape[:2]
    scale = 1.5
    scaled = np.array(Image.fromarray(image).resize((int(w * scale), int(h * scale)), Image.LANCZOS))
    locs_scaled = face_recognition.face_locations(scaled, number_of_times_to_upsample=1, model='hog')
    for (top, right, bottom, left) in locs_scaled:
        orig_loc = (
            int(top / scale),
            int(right / scale),
            int(bottom / scale),
            int(left / scale)
        )
        if not _is_duplicate(orig_loc, all_locations):
            all_locations.add(orig_loc)

    # Pasada 3: imagen escalada x2.0 (para caras muy pequeñas)
    if upsample >= 2:
        scale2 = 2.0
        scaled2 = np.array(Image.fromarray(image).resize((int(w * scale2), int(h * scale2)), Image.LANCZOS))
        locs_scaled2 = face_recognition.face_locations(scaled2, number_of_times_to_upsample=1, model='hog')
        for (top, right, bottom, left) in locs_scaled2:
            orig_loc = (
                int(top / scale2),
                int(right / scale2),
                int(bottom / scale2),
                int(left / scale2)
            )
            if not _is_duplicate(orig_loc, all_locations):
                all_locations.add(orig_loc)

    return list(all_locations)


def _is_duplicate(new_loc, existing_locations, iou_threshold=0.4):
    """Verifica si una detección ya existe (basado en IoU - Intersection over Union)."""
    new_top, new_right, new_bottom, new_left = new_loc
    for (top, right, bottom, left) in existing_locations:
        inter_top = max(new_top, top)
        inter_left = max(new_left, left)
        inter_bottom = min(new_bottom, bottom)
        inter_right = min(new_right, right)

        if inter_bottom <= inter_top or inter_right <= inter_left:
            continue

        inter_area = (inter_bottom - inter_top) * (inter_right - inter_left)
        area1 = (new_bottom - new_top) * (new_right - new_left)
        area2 = (bottom - top) * (right - left)
        union_area = area1 + area2 - inter_area

        if union_area > 0 and inter_area / union_area > iou_threshold:
            return True
    return False


# ─────────────────────────────────────────────
# Gestión de miembros
# ─────────────────────────────────────────────

def load_members():
    """Escanea la carpeta miembros/ y calcula encodings faciales para cada uno."""
    global members_cache
    members_cache = {}

    if not MEMBERS_DIR.exists():
        MEMBERS_DIR.mkdir(parents=True, exist_ok=True)
        return

    loaded = 0
    errors = []
    for file_path in MEMBERS_DIR.iterdir():
        if not file_path.is_file():
            continue
        if file_path.suffix.lower().lstrip('.') not in ALLOWED_EXTENSIONS:
            continue

        name = file_path.stem  # nombre sin extensión
        try:
            img = Image.open(file_path).convert('RGB')
            img_array = np.array(img)
            encodings = face_recognition.face_encodings(img_array)
            if len(encodings) > 0:
                members_cache[name] = encodings[0]
                loaded += 1
            else:
                errors.append(f'{name}: no se detectó rostro')
        except Exception as e:
            errors.append(f'{name}: {str(e)}')

    print(f'[Miembros] Cargados: {loaded} | Errores: {len(errors)}')
    for err in errors:
        print(f'  - {err}')


# ─────────────────────────────────────────────
# Extracción de metadatos EXIF
# ─────────────────────────────────────────────

def extract_photo_date(file_storage):
    """Extrae la fecha de captura: primero EXIF, luego nombre del archivo.
    Retorna (fecha_str, fuente) donde fuente es 'exif', 'nombre' o None."""
    # 1. Intentar EXIF
    try:
        file_storage.seek(0)
        img = Image.open(file_storage)

        exif_data = None
        if hasattr(img, '_getexif') and img._getexif():
            exif_data = img._getexif()

        if not exif_data:
            exif_obj = img.getexif()
            if exif_obj:
                exif_data = dict(exif_obj)

        if exif_data:
            date_value = None
            for tag_id in [36867, 36868, 306]:
                if tag_id in exif_data and exif_data[tag_id]:
                    date_value = exif_data[tag_id]
                    break

            if date_value and isinstance(date_value, str):
                for fmt in ['%Y:%m:%d %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y:%m:%d']:
                    try:
                        dt = datetime.strptime(date_value.strip(), fmt)
                        file_storage.seek(0)
                        print(f'[Fecha] Encontrada en EXIF: {dt.strftime("%Y-%m-%d")}')
                        return dt.strftime('%Y-%m-%d'), 'exif'
                    except ValueError:
                        continue

    except Exception as e:
        print(f'[Fecha] Error extrayendo EXIF: {e}')

    # 2. Intentar extraer fecha del nombre del archivo
    file_storage.seek(0)
    filename = file_storage.filename or ''
    date_from_name = extract_date_from_filename(filename)
    if date_from_name:
        print(f'[Fecha] Encontrada en nombre de archivo "{filename}": {date_from_name}')
        return date_from_name, 'nombre'

    file_storage.seek(0)
    return None, None


def extract_date_from_filename(filename):
    """Extrae una fecha del nombre de archivo. Soporta formatos comunes como
    'WhatsApp Image 2026-02-05 at 8.43.17 PM', 'IMG_20260205_...', '2026-02-05 foto', etc."""
    if not filename:
        return None

    # Patrón: YYYY-MM-DD
    match = re.search(r'(\d{4})-(\d{2})-(\d{2})', filename)
    if match:
        try:
            dt = datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            pass

    # Patrón: YYYYMMDD (ej: IMG_20260205_)
    match = re.search(r'(\d{4})(\d{2})(\d{2})', filename)
    if match:
        try:
            dt = datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
            if 2000 <= dt.year <= 2100:
                return dt.strftime('%Y-%m-%d')
        except ValueError:
            pass

    # Patrón: DD-MM-YYYY o DD/MM/YYYY
    match = re.search(r'(\d{2})[-/](\d{2})[-/](\d{4})', filename)
    if match:
        try:
            dt = datetime(int(match.group(3)), int(match.group(2)), int(match.group(1)))
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            pass

    return None


# ─────────────────────────────────────────────
# Almacenamiento de fotos y duplicados
# ─────────────────────────────────────────────

def _ensure_training_dir():
    """Crea la carpeta entrenamientos/ si no existe."""
    TRAINING_DIR.mkdir(parents=True, exist_ok=True)


def load_photo_index():
    """Carga el índice de fotos procesadas."""
    if INDEX_PATH.exists():
        with open(INDEX_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def save_photo_index(index):
    """Guarda el índice de fotos procesadas."""
    _ensure_training_dir()
    with open(INDEX_PATH, 'w', encoding='utf-8') as f:
        json.dump(index, f, ensure_ascii=False, indent=2)


def compute_file_hash(file_storage):
    """Calcula SHA256 de un archivo."""
    file_storage.seek(0)
    h = hashlib.sha256()
    while True:
        chunk = file_storage.read(8192)
        if not chunk:
            break
        h.update(chunk)
    file_storage.seek(0)
    return h.hexdigest()


def check_duplicate(file_storage, index):
    """Verifica si una foto ya fue procesada. Retorna (es_duplicado, fecha_existente)."""
    file_hash = compute_file_hash(file_storage)
    for date_str, date_data in index.items():
        for photo in date_data.get('photos', []):
            if photo.get('hash') == file_hash:
                return True, date_str, file_hash
    return False, None, file_hash


def save_training_photo(file_storage, date_str, file_hash):
    """Guarda una foto de entrenamiento en entrenamientos/<fecha>/."""
    date_dir = TRAINING_DIR / date_str
    date_dir.mkdir(parents=True, exist_ok=True)
    filename = file_storage.filename or f'foto_{datetime.now().strftime("%H%M%S")}.jpg'
    # Evitar sobreescribir
    dest = date_dir / filename
    counter = 1
    while dest.exists():
        stem = Path(filename).stem
        suffix = Path(filename).suffix
        dest = date_dir / f'{stem}_{counter}{suffix}'
        counter += 1
    file_storage.seek(0)
    dest.write_bytes(file_storage.read())
    file_storage.seek(0)
    return dest.name


def register_photo_in_index(index, date_str, filename, file_hash):
    """Registra una foto en el índice."""
    if date_str not in index:
        index[date_str] = {'photos': []}
    index[date_str]['photos'].append({
        'filename': filename,
        'hash': file_hash,
        'timestamp': datetime.now().isoformat()
    })


# ─────────────────────────────────────────────
# Generación de fotos anotadas por miembro
# ─────────────────────────────────────────────

def _expand_box(top, right, bottom, left, img_h, img_w, factor=0.3):
    """Expande un bounding box por un factor, sin salirse de la imagen."""
    h = bottom - top
    w = right - left
    dh = int(h * factor)
    dw = int(w * factor)
    return (
        max(0, top - dh),
        min(img_w, right + dw),
        min(img_h, bottom + dh),
        max(0, left - dw)
    )


def generate_annotated_photo(photo_image, face_locations, member_name, member_face_idx, date_str):
    """
    Genera una foto grupal anotada con solo un miembro resaltado.
    - photo_image: numpy array RGB
    - face_locations: lista de (top, right, bottom, left)
    - member_name: nombre del miembro a resaltar
    - member_face_idx: índice de la cara que matcheó
    - date_str: fecha para la carpeta
    Retorna el path al archivo guardado.
    """
    img = Image.fromarray(photo_image)
    draw = ImageDraw.Draw(img)
    img_w, img_h = img.size

    # Intentar cargar una fuente con buen tamaño
    font_size = max(16, int(img_w * 0.02))
    try:
        font = ImageFont.truetype("arial.ttf", font_size)
    except (IOError, OSError):
        try:
            font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", font_size)
        except (IOError, OSError):
            font = ImageFont.load_default()

    # Dibujar solo la cara del miembro
    if member_face_idx < len(face_locations):
        top, right, bottom, left = face_locations[member_face_idx]
        et, er, eb, el = _expand_box(top, right, bottom, left, img_h, img_w)

        # Cuadro verde
        line_w = max(3, int(img_w * 0.004))
        for i in range(line_w):
            draw.rectangle([el - i, et - i, er + i, eb + i], outline='#00e676')

        # Etiqueta con nombre
        bbox = font.getbbox(member_name)
        text_w = bbox[2] - bbox[0]
        text_h = bbox[3] - bbox[1]
        label_y = max(0, et - text_h - 8)
        draw.rectangle([el, label_y, el + text_w + 12, label_y + text_h + 6], fill='#00e676')
        draw.text((el + 6, label_y + 2), member_name, fill='white', font=font)

    # Redimensionar a ~800px de ancho
    if img_w > 800:
        ratio = 800 / img_w
        img = img.resize((800, int(img_h * ratio)), Image.LANCZOS)

    # Guardar
    annotated_dir = TRAINING_DIR / date_str / 'anotadas'
    annotated_dir.mkdir(parents=True, exist_ok=True)
    safe_name = re.sub(r'[<>:"/\\|?*]', '_', member_name)
    out_path = annotated_dir / f'{safe_name}.jpg'
    img.save(str(out_path), 'JPEG', quality=85)
    return out_path


# ─────────────────────────────────────────────
# Gestión del Excel de asistencias
# ─────────────────────────────────────────────

def _style_header(ws, row, col):
    """Aplica estilo a una celda de encabezado."""
    cell = ws.cell(row=row, column=col)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')


def _col_letter(col_idx):
    """Convierte índice de columna (1-based) a letra(s) de Excel."""
    from openpyxl.utils import get_column_letter
    return get_column_letter(col_idx)


def update_attendance_excel(date_str, present_members, annotated_paths=None):
    """
    Actualiza el Excel de asistencias con formato de 2 filas por miembro.
    - Fila de datos: Nombre | Total | Mes | ✓ por fecha
    - Fila de fotos: imágenes anotadas debajo de cada ✓
    - annotated_paths: dict {member_name: Path} con fotos anotadas
    """
    if annotated_paths is None:
        annotated_paths = {}

    # Columnas fijas: A=Miembro, B=Total, C=Mes actual
    FIRST_DATE_COL = 4  # Columna D en adelante son fechas

    if EXCEL_PATH.exists():
        wb = load_workbook(str(EXCEL_PATH))
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Asistencias'
        ws.cell(row=1, column=1, value='Miembro')
        ws.cell(row=1, column=2, value='Total')
        ws.cell(row=1, column=3, value='Mes')
        for c in range(1, 4):
            _style_header(ws, 1, c)
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 8

    # Leer fechas existentes (fila 1, desde col FIRST_DATE_COL)
    date_columns = {}
    for col in range(FIRST_DATE_COL, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            date_columns[str(val)] = col

    # Agregar columna de fecha si es nueva
    if date_str not in date_columns:
        new_col = max(ws.max_column + 1, FIRST_DATE_COL)
        ws.cell(row=1, column=new_col, value=date_str)
        _style_header(ws, 1, new_col)
        ws.column_dimensions[_col_letter(new_col)].width = 18
        date_columns[date_str] = new_col

    date_col = date_columns[date_str]

    # Leer miembros existentes (fila de datos = filas pares a partir de 2: 2,4,6...)
    # Formato: fila N = datos, fila N+1 = fotos
    member_data_rows = {}  # nombre -> fila de datos
    row = 2
    while row <= ws.max_row:
        val = ws.cell(row=row, column=1).value
        if val and str(val).strip():
            member_data_rows[str(val).strip()] = row
            row += 2  # saltar fila de fotos
        else:
            row += 1

    # Agregar miembros nuevos
    all_members = sorted(set(list(members_cache.keys()) + list(member_data_rows.keys())))
    for member_name in all_members:
        if member_name not in member_data_rows:
            new_row = ws.max_row + 1 if ws.max_row >= 2 else 2
            # Asegurar que new_row es par (filas de datos en posiciones pares: 2,4,6...)
            if (new_row - 2) % 2 != 0:
                new_row += 1
            ws.cell(row=new_row, column=1, value=member_name)
            ws.cell(row=new_row, column=1).font = Font(bold=True)
            ws.cell(row=new_row, column=1).alignment = Alignment(vertical='center')
            # Fila de fotos (debajo)
            ws.cell(row=new_row + 1, column=1, value='')
            member_data_rows[member_name] = new_row

    # Marcar asistencia y insertar fotos
    current_month = datetime.now().strftime('%Y-%m')

    for member_name in all_members:
        data_row = member_data_rows[member_name]
        photo_row = data_row + 1

        if member_name in present_members:
            cell = ws.cell(row=data_row, column=date_col, value='✓')
            cell.font = Font(color='008000', bold=True, size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Insertar foto anotada si existe
            if member_name in annotated_paths:
                img_path = annotated_paths[member_name]
                if img_path.exists():
                    try:
                        xl_img = XlImage(str(img_path))
                        # Tamaño máximo en píxeles que cabe en la celda
                        # Col width 18 chars ≈ 126px, row height 120pt ≈ 160px
                        max_w, max_h = 120, 150
                        orig_w = xl_img.width
                        orig_h = xl_img.height
                        if orig_w > 0 and orig_h > 0:
                            scale = min(max_w / orig_w, max_h / orig_h)
                            xl_img.width = int(orig_w * scale)
                            xl_img.height = int(orig_h * scale)
                        anchor = f'{_col_letter(date_col)}{photo_row}'
                        xl_img.anchor = anchor
                        ws.add_image(xl_img)
                    except Exception as e:
                        print(f'[Excel] Error insertando imagen para {member_name}: {e}')
        else:
            existing = ws.cell(row=data_row, column=date_col).value
            if not existing:
                ws.cell(row=data_row, column=date_col, value='')

        # Altura de fila de fotos: suficiente para la imagen
        ws.row_dimensions[photo_row].height = 115

        # Recalcular totales
        total = 0
        month_total = 0
        for d_str, d_col in date_columns.items():
            val = ws.cell(row=data_row, column=d_col).value
            if val == '✓':
                total += 1
                if d_str.startswith(current_month):
                    month_total += 1

        total_cell = ws.cell(row=data_row, column=2, value=total)
        total_cell.font = Font(bold=True, color='0066CC', size=12)
        total_cell.alignment = Alignment(horizontal='center', vertical='center')

        month_cell = ws.cell(row=data_row, column=3, value=month_total)
        month_cell.font = Font(bold=True, color='008080', size=12)
        month_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ajustar ancho de columna A
    if all_members:
        ws.column_dimensions['A'].width = max(22, max((len(n) + 2 for n in all_members), default=22))

    wb.save(str(EXCEL_PATH))
    return True


def get_attendance_summary():
    """Lee el Excel y retorna un resumen de asistencias para el frontend."""
    if not EXCEL_PATH.exists():
        return {'dates': [], 'members': []}

    wb = load_workbook(str(EXCEL_PATH))
    ws = wb.active

    FIRST_DATE_COL = 4

    dates = []
    for col in range(FIRST_DATE_COL, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            dates.append(str(val))

    members = []
    row = 2
    while row <= ws.max_row:
        name = ws.cell(row=row, column=1).value
        if not name or not str(name).strip():
            row += 1
            continue
        name = str(name).strip()
        attendance = {}
        total = 0
        current_month = datetime.now().strftime('%Y-%m')
        month_total = 0
        for col_idx, date in enumerate(dates, start=FIRST_DATE_COL):
            val = ws.cell(row=row, column=col_idx).value
            present = val == '✓'
            attendance[date] = present
            if present:
                total += 1
                if date.startswith(current_month):
                    month_total += 1
        members.append({
            'name': name,
            'attendance': attendance,
            'total': total,
            'month_total': month_total
        })
        row += 2  # saltar fila de fotos

    return {'dates': dates, 'members': members}


# ─────────────────────────────────────────────
# Rutas
# ─────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/members', methods=['GET'])
def get_members():
    """Retorna la lista de miembros cargados."""
    return jsonify({
        'count': len(members_cache),
        'members': sorted(members_cache.keys())
    })


@app.route('/member-photo/<name>', methods=['GET'])
def serve_member_photo(name):
    """Sirve la foto de referencia de un miembro."""
    for ext in ('jpg', 'jpeg', 'png', 'webp', 'bmp'):
        path = MEMBERS_DIR / f'{name}.{ext}'
        if path.exists():
            return send_file(str(path))
    return jsonify({'error': 'Foto no encontrada.'}), 404


@app.route('/reload-members', methods=['POST'])
def reload_members():
    """Recarga la base de miembros desde la carpeta."""
    load_members()
    return jsonify({
        'count': len(members_cache),
        'members': sorted(members_cache.keys()),
        'message': f'Se cargaron {len(members_cache)} miembros.'
    })


def _find_first_image(directory):
    """Busca la primera imagen en un directorio."""
    if directory.exists():
        for f in sorted(directory.iterdir()):
            if f.suffix.lower() in ('.jpg', '.jpeg', '.png', '.webp', '.bmp'):
                return f
    return None


def _load_demo_members(demo_members_dir):
    """Carga encodings de miembros de un directorio de demo."""
    demo_cache = {}
    if not demo_members_dir.exists():
        return demo_cache
    for img_path in sorted(demo_members_dir.iterdir()):
        if img_path.suffix.lower() not in ('.jpg', '.jpeg', '.png', '.webp', '.bmp'):
            continue
        name = img_path.stem
        try:
            img = face_recognition.load_image_file(str(img_path))
            encs = face_recognition.face_encodings(img)
            if encs:
                demo_cache[name] = encs[0]
        except Exception:
            pass
    return demo_cache


@app.route('/demos', methods=['GET'])
def list_demos():
    """Lista los demos disponibles en demos/."""
    demos = []
    if DEMOS_DIR.exists():
        for d in sorted(DEMOS_DIR.iterdir()):
            if not d.is_dir():
                continue
            members_dir = d / 'miembros'
            grupal_dir = d / 'grupal'
            group_photo = _find_first_image(grupal_dir)
            member_count = len([f for f in members_dir.iterdir()
                               if f.suffix.lower() in ('.jpg', '.jpeg', '.png', '.webp', '.bmp')]) if members_dir.exists() else 0
            if group_photo and member_count > 0:
                label = d.name.replace('_', ' ').title()
                demos.append({
                    'id': d.name,
                    'label': label,
                    'member_count': member_count,
                    'group_photo_url': f'/demo-photo/{d.name}'
                })
    return jsonify({'demos': demos})


@app.route('/demo-photo/<demo_id>', methods=['GET'])
def serve_demo_photo(demo_id):
    """Sirve la foto grupal de un demo específico."""
    grupal_dir = DEMOS_DIR / demo_id / 'grupal'
    photo = _find_first_image(grupal_dir)
    if photo:
        return send_file(str(photo))
    return jsonify({'error': 'No hay foto de demo.'}), 404


@app.route('/demo-member-photo/<demo_id>/<name>', methods=['GET'])
def serve_demo_member_photo(demo_id, name):
    """Sirve la foto de referencia de un miembro de un demo."""
    members_dir = DEMOS_DIR / demo_id / 'miembros'
    for ext in ('jpg', 'jpeg', 'png', 'webp', 'bmp'):
        path = members_dir / f'{name}.{ext}'
        if path.exists():
            return send_file(str(path))
    return jsonify({'error': 'Foto no encontrada.'}), 404


@app.route('/demo-attendance', methods=['POST'])
def demo_attendance():
    """Procesa la foto grupal de un demo contra sus propios miembros."""
    demo_id = request.form.get('demo_id', '').strip()
    if not demo_id:
        return jsonify({'error': 'Se requiere un demo_id.'}), 400

    demo_dir = DEMOS_DIR / demo_id
    if not demo_dir.exists():
        return jsonify({'error': f'Demo "{demo_id}" no encontrado.'}), 404

    grupal_dir = demo_dir / 'grupal'
    members_dir = demo_dir / 'miembros'
    demo_photo = _find_first_image(grupal_dir)
    if not demo_photo:
        return jsonify({'error': 'No hay foto grupal en este demo.'}), 404

    # Cargar miembros del demo
    demo_members = _load_demo_members(members_dir)
    if len(demo_members) == 0:
        return jsonify({'error': 'No hay miembros cargados en este demo.'}), 400

    try:
        demo_image = face_recognition.load_image_file(str(demo_photo))

        upsample = int(request.form.get('upsample', 2))
        tolerance = float(request.form.get('tolerance', 0.75))
        use_multiscale = request.form.get('model', 'hog') == 'multiscale'

        if use_multiscale:
            face_locations = detect_faces_multiscale(demo_image, upsample=upsample)
        else:
            face_locations = face_recognition.face_locations(demo_image, number_of_times_to_upsample=upsample, model='hog')

        encodings = face_recognition.face_encodings(demo_image, face_locations)

        if len(encodings) == 0:
            return jsonify({'error': 'No se detectaron rostros en la foto de demo.'}), 400

        member_names = sorted(demo_members.keys())
        member_encodings_list = [demo_members[name] for name in member_names]

        # Construir matriz de distancias (miembros x caras)
        dist_matrix = np.zeros((len(member_names), len(encodings)))
        for i, member_enc in enumerate(member_encodings_list):
            dist_matrix[i] = face_recognition.face_distance(encodings, member_enc)

        # Asignación exclusiva greedy
        assigned_faces = set()
        assigned_members = {}

        pairs = []
        for mi in range(len(member_names)):
            for fi in range(len(encodings)):
                pairs.append((dist_matrix[mi, fi], mi, fi))
        pairs.sort(key=lambda x: x[0])

        for dist, mi, fi in pairs:
            if mi in assigned_members or fi in assigned_faces:
                continue
            if dist <= tolerance:
                assigned_members[mi] = (fi, dist)
                assigned_faces.add(fi)

        present_members = []
        member_results = []
        demo_ann_tag = f'_demo_{demo_id}'

        for i, name in enumerate(member_names):
            if i in assigned_members:
                face_idx, best_distance = assigned_members[i]
                confidence = round((1 - best_distance) * 100, 1)
                present_members.append(name)
                generate_annotated_photo(
                    demo_image, face_locations,
                    name, face_idx, demo_ann_tag
                )
                safe = re.sub(r'[<>:"/\\|?*]', '_', name)
                result_entry = {
                    'name': name,
                    'present': True,
                    'confidence': confidence,
                    'annotated_url': f'/training-photo/{demo_ann_tag}/anotadas/{safe}.jpg',
                    'member_photo_url': f'/demo-member-photo/{demo_id}/{name}'
                }
            else:
                best_distance = float(np.min(dist_matrix[i]))
                confidence = round((1 - best_distance) * 100, 1)
                result_entry = {
                    'name': name,
                    'present': False,
                    'confidence': confidence
                }
            member_results.append(result_entry)

        demo_label = demo_id.replace('_', ' ').title()
        return jsonify({
            'date': f'DEMO - {demo_label}',
            'date_source': 'demo',
            'photos_processed': 1,
            'total_faces_detected': len(encodings),
            'total_members': len(member_names),
            'present_count': len(present_members),
            'absent_count': len(member_names) - len(present_members),
            'results': member_results,
            'is_demo': True,
            'demo_id': demo_id,
            'demo_photo': demo_photo.name
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Error en demo: {str(e)}'}), 500


@app.route('/attendance', methods=['POST'])
def register_attendance():
    """Procesa una o varias fotos grupales y registra asistencia."""
    photo_files = request.files.getlist('photos')
    if not photo_files or all(not f.filename for f in photo_files):
        return jsonify({'error': 'Se requiere al menos una foto grupal.'}), 400

    # Filtrar archivos válidos
    valid_files = [f for f in photo_files if f.filename and allowed_file(f.filename)]
    if not valid_files:
        return jsonify({'error': 'Formato de archivo no permitido.'}), 400

    if len(members_cache) == 0:
        return jsonify({'error': 'No hay miembros cargados. Agrega fotos a la carpeta miembros/ y recarga.'}), 400

    # Detección de duplicados
    photo_index = load_photo_index()
    duplicates = []
    new_files = []
    file_hashes = []
    for f in valid_files:
        is_dup, dup_date, f_hash = check_duplicate(f, photo_index)
        if is_dup:
            duplicates.append({'filename': f.filename, 'date': dup_date})
        else:
            new_files.append(f)
            file_hashes.append(f_hash)

    if not new_files:
        dup_info = '; '.join([f'"{d["filename"]}" ya registrada en {d["date"]}' for d in duplicates])
        return jsonify({'error': f'Todas las fotos ya fueron procesadas. {dup_info}'}), 400

    # Extraer fecha de la primera foto nueva (EXIF o nombre de archivo)
    auto_date, auto_source = extract_photo_date(new_files[0])
    manual_date = request.form.get('date', '').strip()
    if manual_date:
        date_str = manual_date
        date_source = 'manual'
    elif auto_date:
        date_str = auto_date
        date_source = auto_source
    else:
        date_str = datetime.now().strftime('%Y-%m-%d')
        date_source = 'hoy'

    # Parámetros
    upsample = int(request.form.get('upsample', 2))
    tolerance = float(request.form.get('tolerance', 0.75))
    use_multiscale = request.form.get('model', 'hog') == 'multiscale'

    try:
        # Procesar cada foto: guardar, detectar rostros, guardar metadatos por foto
        # Necesitamos saber de qué foto vino cada encoding para generar la foto anotada correcta
        photo_data = []  # lista de {image, locations, encodings, encoding_offset}
        all_face_encodings = []
        total_faces = 0

        for photo_file, f_hash in zip(new_files, file_hashes):
            # Guardar foto en disco
            saved_name = save_training_photo(photo_file, date_str, f_hash)
            register_photo_in_index(photo_index, date_str, saved_name, f_hash)

            photo_file.seek(0)
            photo_image = load_image_from_upload(photo_file)

            if use_multiscale:
                face_locations = detect_faces_multiscale(photo_image, upsample=upsample)
            else:
                face_locations = face_recognition.face_locations(photo_image, number_of_times_to_upsample=upsample, model='hog')

            encodings = face_recognition.face_encodings(photo_image, face_locations)

            photo_data.append({
                'image': photo_image,
                'locations': face_locations,
                'encodings': encodings,
                'encoding_offset': total_faces
            })

            all_face_encodings.extend(encodings)
            total_faces += len(encodings)

        # Guardar índice actualizado
        save_photo_index(photo_index)

        if total_faces == 0:
            return jsonify({
                'error': f'No se detectaron rostros en {len(new_files)} foto(s).',
                'date': date_str,
                'date_source': date_source
            }), 400

        # Comparar cada miembro contra todos los rostros de todas las fotos
        member_names = sorted(members_cache.keys())
        member_encodings_list = [members_cache[name] for name in member_names]

        # Construir matriz de distancias (miembros x caras)
        all_enc_array = all_face_encodings
        dist_matrix = np.zeros((len(member_names), total_faces))
        for i, member_enc in enumerate(member_encodings_list):
            dist_matrix[i] = face_recognition.face_distance(all_enc_array, member_enc)

        # Asignación exclusiva: cada cara solo puede asignarse a un miembro
        assigned_faces = set()
        assigned_members = {}  # {member_idx: (face_idx, distance)}

        pairs = []
        for mi in range(len(member_names)):
            for fi in range(total_faces):
                pairs.append((dist_matrix[mi, fi], mi, fi))
        pairs.sort(key=lambda x: x[0])

        for dist, mi, fi in pairs:
            if mi in assigned_members or fi in assigned_faces:
                continue
            if dist <= tolerance:
                assigned_members[mi] = (fi, dist)
                assigned_faces.add(fi)

        present_members = []
        member_results = []
        annotated_paths = {}  # {member_name: Path}

        for i, name in enumerate(member_names):
            if i in assigned_members:
                best_idx, best_distance = assigned_members[i]
                confidence = round((1 - best_distance) * 100, 1)
                present_members.append(name)

                # Encontrar a qué foto corresponde el best_idx
                for pd in photo_data:
                    offset = pd['encoding_offset']
                    count = len(pd['encodings'])
                    if offset <= best_idx < offset + count:
                        local_idx = best_idx - offset
                        ann_path = generate_annotated_photo(
                            pd['image'], pd['locations'],
                            name, local_idx, date_str
                        )
                        annotated_paths[name] = ann_path
                        break

                safe = re.sub(r'[<>:"/\\|?*]', '_', name)
                result_entry = {
                    'name': name,
                    'present': True,
                    'confidence': confidence,
                    'annotated_url': f'/training-photo/{date_str}/anotadas/{safe}.jpg',
                    'member_photo_url': f'/member-photo/{name}'
                }
            else:
                best_distance = float(np.min(dist_matrix[i]))
                confidence = round((1 - best_distance) * 100, 1)
                result_entry = {
                    'name': name,
                    'present': False,
                    'confidence': confidence
                }
            member_results.append(result_entry)

        # Actualizar Excel con fotos anotadas
        update_attendance_excel(date_str, present_members, annotated_paths)

        response = {
            'date': date_str,
            'date_source': date_source,
            'photos_processed': len(new_files),
            'total_faces_detected': total_faces,
            'total_members': len(member_names),
            'present_count': len(present_members),
            'absent_count': len(member_names) - len(present_members),
            'results': member_results
        }
        if duplicates:
            response['duplicates_skipped'] = duplicates

        return jsonify(response)

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Error durante el análisis: {str(e)}'}), 500


@app.route('/attendance-summary', methods=['GET'])
def attendance_summary():
    """Retorna el resumen completo de asistencias."""
    try:
        summary = get_attendance_summary()
        return jsonify(summary)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/delete-attendance', methods=['POST'])
def delete_attendance():
    """Elimina una columna de fecha del Excel de asistencias."""
    date_str = request.json.get('date', '').strip() if request.is_json else ''
    if not date_str:
        return jsonify({'error': 'Se requiere una fecha para eliminar.'}), 400

    if not EXCEL_PATH.exists():
        return jsonify({'error': 'No hay archivo de asistencias.'}), 404

    try:
        wb = load_workbook(str(EXCEL_PATH))
        ws = wb.active

        # Buscar la columna con esa fecha (desde col 4 en adelante)
        target_col = None
        for col in range(4, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if str(val) == date_str:
                target_col = col
                break

        if not target_col:
            return jsonify({'error': f'No se encontró la fecha {date_str} en el registro.'}), 404

        ws.delete_cols(target_col)
        wb.save(str(EXCEL_PATH))

        return jsonify({'message': f'Asistencia del {date_str} eliminada correctamente.'})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Error al eliminar: {str(e)}'}), 500


@app.route('/download-excel', methods=['GET'])
def download_excel():
    """Descarga el archivo Excel de asistencias."""
    if not EXCEL_PATH.exists():
        return jsonify({'error': 'No hay archivo de asistencias aún.'}), 404
    return send_file(str(EXCEL_PATH), as_attachment=True, download_name='asistencias.xlsx')


@app.route('/training-photos', methods=['GET'])
def get_training_photos():
    """Retorna la lista de fotos de entrenamiento agrupadas por fecha."""
    index = load_photo_index()
    result = {}
    for date_str, date_data in sorted(index.items(), reverse=True):
        photos = date_data.get('photos', [])
        result[date_str] = [p['filename'] for p in photos]
    return jsonify(result)


@app.route('/training-photo/<date_str>/<path:filename>', methods=['GET'])
def get_training_photo(date_str, filename):
    """Sirve una foto de entrenamiento o anotada."""
    file_path = TRAINING_DIR / date_str / filename
    if not file_path.exists():
        # Intentar en subcarpeta anotadas
        file_path = TRAINING_DIR / date_str / 'anotadas' / filename
    if not file_path.exists():
        return jsonify({'error': 'Foto no encontrada.'}), 404
    return send_file(str(file_path))


@app.route('/delete-training-photo', methods=['POST'])
def delete_training_photo():
    """Elimina una foto de entrenamiento del disco y del índice."""
    data = request.get_json() if request.is_json else {}
    date_str = data.get('date', '').strip()
    filename = data.get('filename', '').strip()
    if not date_str or not filename:
        return jsonify({'error': 'Se requiere fecha y nombre de archivo.'}), 400

    # Eliminar del índice
    index = load_photo_index()
    if date_str in index:
        photos = index[date_str].get('photos', [])
        index[date_str]['photos'] = [p for p in photos if p['filename'] != filename]
        if not index[date_str]['photos']:
            del index[date_str]
        save_photo_index(index)

    # Eliminar archivo del disco
    file_path = TRAINING_DIR / date_str / filename
    if file_path.exists():
        file_path.unlink()

    # Eliminar anotadas asociadas
    annotated_dir = TRAINING_DIR / date_str / 'anotadas'
    if annotated_dir.exists():
        import shutil
        shutil.rmtree(str(annotated_dir), ignore_errors=True)

    # Si la carpeta de fecha quedó vacía, eliminarla
    date_dir = TRAINING_DIR / date_str
    if date_dir.exists() and not any(date_dir.iterdir()):
        date_dir.rmdir()

    return jsonify({'message': f'Foto "{filename}" del {date_str} eliminada.'})


@app.route('/upload-excel', methods=['POST'])
def upload_excel():
    """Importa un Excel existente y hace merge con el actual."""
    if 'file' not in request.files:
        return jsonify({'error': 'Se requiere un archivo Excel.'}), 400

    excel_file = request.files['file']
    if not excel_file.filename or not excel_file.filename.endswith('.xlsx'):
        return jsonify({'error': 'El archivo debe ser .xlsx'}), 400

    try:
        # Leer el Excel importado
        imported_wb = load_workbook(excel_file)
        imported_ws = imported_wb.active

        # Detectar formato: buscar columnas de fecha
        # Intentar formato nuevo (col 4+) o viejo (col 2+)
        imported_dates = {}
        first_date_col = None
        for col in range(2, imported_ws.max_column + 1):
            val = imported_ws.cell(row=1, column=col).value
            if val and re.match(r'\d{4}-\d{2}-\d{2}', str(val)):
                if first_date_col is None:
                    first_date_col = col
                imported_dates[str(val)] = col

        if not imported_dates:
            return jsonify({'error': 'No se encontraron fechas de asistencia en el Excel.'}), 400

        # Leer miembros y asistencias del importado
        imported_members = {}  # {nombre: {fecha: bool}}
        row = 2
        while row <= imported_ws.max_row:
            name = imported_ws.cell(row=row, column=1).value
            if name and str(name).strip() and str(name).strip() not in ('Total', 'Mes', 'Miembro'):
                name = str(name).strip()
                attendance = {}
                for date_str, col in imported_dates.items():
                    val = imported_ws.cell(row=row, column=col).value
                    if val == '✓':
                        attendance[date_str] = True
                if attendance:
                    imported_members[name] = attendance
            row += 1

        # Merge: para cada miembro/fecha importado, registrar si no existe
        dates_added = set()
        marks_added = 0
        for name, attendance in imported_members.items():
            for date_str, present in attendance.items():
                if present:
                    # Asegurar que el miembro existe en cache (o al menos en Excel)
                    update_attendance_excel(date_str, [name])
                    dates_added.add(date_str)
                    marks_added += 1

        return jsonify({
            'message': f'Importación completada: {len(imported_members)} miembros, {len(dates_added)} fechas, {marks_added} marcas de asistencia.',
            'members_imported': list(imported_members.keys()),
            'dates_imported': sorted(dates_added)
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Error al importar: {str(e)}'}), 500


@app.route('/compare', methods=['POST'])
def compare_faces():
    if 'reference' not in request.files or 'target' not in request.files:
        return jsonify({'error': 'Se requieren ambas imágenes: referencia y objetivo.'}), 400

    ref_file = request.files['reference']
    target_file = request.files['target']

    if not ref_file.filename or not target_file.filename:
        return jsonify({'error': 'No se seleccionaron archivos.'}), 400

    if not allowed_file(ref_file.filename) or not allowed_file(target_file.filename):
        return jsonify({'error': 'Formato de archivo no permitido. Use: png, jpg, jpeg, webp, bmp.'}), 400

    try:
        ref_image = load_image_from_upload(ref_file)
        target_image = load_image_from_upload(target_file)
    except Exception as e:
        return jsonify({'error': f'Error al cargar las imágenes: {str(e)}'}), 400

    # Parámetros del frontend
    upsample = int(request.form.get('upsample', 2))
    tolerance = float(request.form.get('tolerance', 0.6))
    use_multiscale = request.form.get('model', 'hog') == 'multiscale'

    try:
        # Obtener encoding de la foto de referencia
        ref_locations = face_recognition.face_locations(ref_image, number_of_times_to_upsample=upsample, model='hog')
        ref_encodings = face_recognition.face_encodings(ref_image, ref_locations)
        if len(ref_encodings) == 0:
            return jsonify({
                'error': 'No se detectó ningún rostro en la foto de referencia. Sube una foto donde se vea claramente tu cara.'
            }), 400

        ref_encoding = ref_encodings[0]

        # Detectar caras en la foto objetivo
        if use_multiscale:
            target_locations = detect_faces_multiscale(target_image, upsample=upsample)
        else:
            target_locations = face_recognition.face_locations(target_image, number_of_times_to_upsample=upsample, model='hog')

        target_encodings = face_recognition.face_encodings(target_image, target_locations)

        if len(target_encodings) == 0:
            return jsonify({
                'found': False,
                'message': 'No se detectaron rostros en la foto objetivo.',
                'total_faces': 0,
                'matches': []
            })

        distances = face_recognition.face_distance(target_encodings, ref_encoding)

        # Solo marcar como match el rostro con menor distancia (mayor confianza)
        best_idx = int(np.argmin(distances))
        best_distance = distances[best_idx]
        best_is_match = bool(best_distance <= tolerance)

        matches = []
        for i, (distance, location) in enumerate(zip(distances, target_locations)):
            top, right, bottom, left = location
            confidence = round((1 - distance) * 100, 1)
            matches.append({
                'face_index': i + 1,
                'is_match': bool(i == best_idx and best_is_match),
                'confidence': confidence,
                'location': {
                    'top': int(top),
                    'right': int(right),
                    'bottom': int(bottom),
                    'left': int(left)
                }
            })

        found = best_is_match
        best_match = matches[best_idx] if found else None

        if found:
            message = f'¡Te encontré! Se detectó tu rostro con {best_match["confidence"]}% de confianza.'
        else:
            message = 'No se encontró tu rostro en la foto objetivo.'

        return jsonify({
            'found': found,
            'message': message,
            'total_faces': len(target_encodings),
            'matches': matches
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Error durante el análisis: {str(e)}'}), 500


# ─────────────────────────────────────────────
# Inicialización
# ─────────────────────────────────────────────

# Cargar miembros al iniciar
load_members()

if __name__ == '__main__':
    app.run(debug=True, port=5000)
